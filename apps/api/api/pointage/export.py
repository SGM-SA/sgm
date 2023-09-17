from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font

from api.pointage.models import PointageEtape


def export_pointages(start_date, end_date):
    """
    Exporte les pointages entre deux dates
    """
    pointages = PointageEtape.objects.exclude(date_fin=None).filter(
        date_debut__range=(start_date, end_date)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Pointages"

    ws.append(
        [
            "Employé",
            "Affaire",
            "Fiche",
            "Etape",
            "Date de début",
            "Date de fin",
            "Durée totale",
        ]
    )

    date_style = NamedStyle(name="datetime", number_format="DD/MM/YYYY HH:MM")

    # seconds to HH:MM:SS
    duree_style = NamedStyle(name="duration", number_format="HH:MM:SS")

    header_font = Font(size=14, bold=True)

    for pointage in pointages:
        ws.append(
            [
                pointage.user.name,
                pointage.etape.fiche.affaire.num_affaire,
                pointage.etape.fiche.titre,
                pointage.etape.num_etape,
                pointage.date_debut.strftime("%d/%m/%Y %H:%M"),
                pointage.date_fin.strftime("%d/%m/%Y %H:%M"),
                seconds_to_hms(pointage.duree()),
            ]
        )
    # Set date style to column E and F
    # not using ws.column_dimensions['E'].style = date_style
    # because it doesn't work, do not appy to header
    for cell in ws["E"]:
        cell.style = date_style

    for cell in ws["F"]:
        cell.style = date_style
    # Set duree style to column G
    for cell in ws["G"]:
        cell.style = duree_style

    # Set header font
    for cell in ws[1]:
        cell.font = header_font

    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20

    response = HttpResponse(content_type="application/ms-excel")
    response[
        "Content-Disposition"
    ] = f'attachment; filename="pointages_{start_date}_{end_date}.xlsx"'
    wb.save(response)

    return response


def seconds_to_hms(seconds):
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return "{:02}:{:02}:{:02}".format(int(hours), int(minutes), int(seconds))
